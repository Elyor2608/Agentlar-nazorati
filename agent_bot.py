"""
🤖 Bot 3.2 - Mijozlar bazasi + Qidiruv + Lokatsiya
"""

import telebot
from telebot import types
from datetime import datetime, date, timedelta, timezone
import json, os, time, math, threading, tempfile

# =============================================
# 🌏 O'ZBEKISTON VAQTI (UTC+5)
# =============================================
UZ_TZ = timezone(timedelta(hours=5))

# =============================================
BOT_TOKEN = "8663649354:AAGDw8eBwEVb5ck7yEbb4hX6Ya3cCZAUxMY"
ADMIN_ID  = 7642574758
# =============================================

PRODUCTS = [
    {"name": "🥪 Sendwich",         "price": 10000},
    {"name": "🧒 Detskiy sendwich", "price": 5000},
    {"name": "🍗 Tovuqli sendwich", "price": 12000},
    {"name": "🍔 Burger",           "price": 15000},
]

ROLES = {
    "admin":       "👑 Admin",
    "supervisor":  "🔷 Supervayzor",
    "manager":     "🔶 Menejer",
    "agent":       "👤 Agent",
}

bot          = telebot.TeleBot(BOT_TOKEN)
DATA_FILE    = "bot_data.json"
SHOPS_FILE   = "shops.json"
CLIENTS_FILE = "clients.json"
sessions     = {}

# =============================================
# 🗄 MA'LUMOTLAR
# =============================================
def load():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {"users": {}, "reports": [], "pending": {}}

def save(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_shops():
    if os.path.exists(SHOPS_FILE):
        with open(SHOPS_FILE, encoding="utf-8") as f:
            return json.load(f)
    return []

def save_shops(shops):
    with open(SHOPS_FILE, "w", encoding="utf-8") as f:
        json.dump(shops, f, ensure_ascii=False, indent=2)

def load_clients():
    if os.path.exists(CLIENTS_FILE):
        try:
            with open(CLIENTS_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"❌ Xato: {e}")
    return []

def save_clients(clients):
    with open(CLIENTS_FILE, "w", encoding="utf-8") as f:
        json.dump(clients, f, ensure_ascii=False, indent=2)

def get_user(uid):
    return load()["users"].get(str(uid))

def is_admin(uid):
    return uid == ADMIN_ID

def now_str():
    return datetime.now(UZ_TZ).strftime("%Y-%m-%d %H:%M")

def today_str():
    return datetime.now(UZ_TZ).strftime("%Y-%m-%d")

def fmt(n):
    return f"{n:,}".replace(",", " ")

def haversine(lat1, lon1, lat2, lon2):
    R = 6371000
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlam/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))

# =============================================
# 🏪 MAGAZINLAR
# =============================================
def find_nearby_shops(lat, lon, radius=100):
    shops = load_shops()
    nearby = []
    for shop in shops:
        dist = haversine(lat, lon, shop["lat"], shop["lon"])
        if dist <= radius:
            shop["distance"] = int(dist)
            nearby.append(shop)
    nearby.sort(key=lambda x: x["distance"])
    return nearby

def add_shop(name, lat, lon):
    shops = load_shops()
    for s in shops:
        if s["name"].lower() == name.lower() and haversine(lat, lon, s["lat"], s["lon"]) < 50:
            return s
    shop = {"id": len(shops) + 1, "name": name, "lat": lat, "lon": lon, "created": now_str()}
    shops.append(shop)
    save_shops(shops)
    return shop

def find_nearby_clients_gps(lat, lon, radius=100):
    """Excel bazasidan va clients.json dan 100m ichidagi klientlarni qaytaradi."""
    results = []
    # 1. clients_cache.json — Excel fayldan yuklangan klientlar
    if os.path.exists("clients_cache.json"):
        try:
            with open("clients_cache.json", encoding="utf-8") as f:
                cache = json.load(f)
            for c in cache:
                clat = c.get("lat")
                clng = c.get("lng")
                if clat and clng:
                    dist = haversine(lat, lon, clat, clng)
                    if dist <= radius:
                        results.append({"name": c["name"], "address": c.get("address", "—"), "distance": int(dist)})
        except Exception:
            pass
    # 2. clients.json — qo'lda kiritilgan klientlar
    for c in load_clients():
        clat = c.get("lat")
        clng = c.get("lng") or c.get("lon")
        if clat and clng:
            dist = haversine(lat, lon, clat, clng)
            if dist <= radius:
                results.append({"name": c["name"], "address": c.get("address", "—"), "distance": int(dist)})
    results.sort(key=lambda x: x["distance"])
    return results

# =============================================
# 👥 MIJOZLAR
# =============================================
def add_client(name, address, phone, lat=None, lon=None):
    clients = load_clients()
    client = {
        "id": len(clients) + 1,
        "name": name,
        "address": address,
        "phone": phone,
        "lat": lat,
        "lon": lon,
        "created": now_str()
    }
    clients.append(client)
    save_clients(clients)
    return client

def search_clients(query):
    clients = load_clients()
    query = query.lower()
    return [c for c in clients if query in c["name"].lower() or query in c.get("address","").lower()]

# =============================================
# KLAVIATURALAR
# =============================================
def main_kb(role):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    if role == "agent":
        kb.add("🛒 Savdoni boshlash")
        kb.add("📊 Mening hisobotlarim")
        kb.add("👥 Mijozlar")
    elif role in ("supervisor", "manager"):
        kb.add("📊 Hisobotlar", "👥 Agentlar")
        kb.add("💰 Sotuv statistikasi", "🏆 Reyting")
        kb.add("🏪 Magazinlar", "👥 Mijozlar")
    return kb

def admin_kb():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add("📊 Hisobotlar", "👥 Foydalanuvchilar")
    kb.add("💰 Sotuv statistikasi", "🏆 Reyting")
    kb.add("🏪 Magazinlar", "👥 Mijozlar")
    kb.add("⏳ Kutayotganlar", "🔄 Hammasini restart")
    kb.add("📤 Kunlik hisobot")
    return kb

def back_kb():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add("⬅️ Orqaga")
    return kb

def location_kb():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    kb.add(types.KeyboardButton("📍 Lokatsiyani yuborish", request_location=True))
    kb.add("⬅️ Orqaga")
    return kb

def products_kb(counts):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    for i, p in enumerate(PRODUCTS):
        cnt = counts.get(str(i), 0)
        label = f"{p['name']} [{cnt} ta]" if cnt > 0 else p['name']
        kb.add(label)
    kb.add("✅ Tayyor")
    kb.add("⬅️ Orqaga")
    return kb

def qty_numpad_kb():
    """Sotuv uchun raqamli klavyatura — faqat tasdiqlash va orqaga"""
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add("✅ Tasdiqlash")
    kb.add("⬅️ Orqaga")
    return kb

def qty_numpad_vozv_kb():
    """Vozvrat uchun raqamli klavyatura — faqat tasdiqlash va orqaga"""
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add("✅ Tasdiqlash")
    kb.add("⬅️ Orqaga")
    return kb

def vozvrat_kb(counts):
    """Vozvrat uchun qizil belgili mahsulotlar paneli"""
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    for i, p in enumerate(PRODUCTS):
        cnt = counts.get(str(i), 0)
        label = f"🔴 {p['name']} [{cnt} ta]" if cnt > 0 else f"🔴 {p['name']}"
        kb.add(label)
    kb.add("✅ Tayyor")
    kb.add("⬅️ Orqaga")
    return kb

# =============================================
# /START
# =============================================
@bot.message_handler(commands=["start"])
def cmd_start(msg):
    uid = msg.from_user.id
    if is_admin(uid):
        data = load()
        if str(uid) not in data["users"]:
            data["users"][str(uid)] = {"name": "Admin", "role": "admin", "registered": now_str(), "approved": True}
            save(data)
        return bot.send_message(uid, "👑 <b>Admin panel</b>\nXush kelibsiz!", parse_mode="HTML", reply_markup=admin_kb())
    user = get_user(uid)
    if user:
        if not user.get("approved"):
            return bot.send_message(uid, "⏳ So'rovingiz hali tasdiqlanmagan.")
        bot.send_message(uid, f"👋 Xush kelibsiz, <b>{user['name']}</b>! ({ROLES.get(user['role'],'')})", parse_mode="HTML", reply_markup=main_kb(user["role"]))
    else:
        bot.send_message(uid, "👋 Salom! Botdan foydalanish uchun <b>Ism Familiyangizni</b> yuboring:", parse_mode="HTML")
        sessions[uid] = {"step": "register_name"}

# =============================================
# RO'YXATDAN O'TISH
# =============================================
@bot.message_handler(func=lambda m: sessions.get(m.from_user.id, {}).get("step") == "register_name")
def register_name(msg):
    uid, name = msg.from_user.id, msg.text.strip()
    sessions[uid] = {"step": "register_role", "name": name}
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add("👤 Agent", "🔷 Supervayzor", "🔶 Menejer")
    bot.send_message(uid, f"Ismingiz: <b>{name}</b>\n\nXodimni tanlang:", parse_mode="HTML", reply_markup=kb)

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id, {}).get("step") == "register_role")
def register_role(msg):
    uid, text = msg.from_user.id, msg.text.strip()
    role_map = {"👤 Agent": "agent", "🔷 Supervayzor": "supervisor", "🔶 Menejer": "manager"}
    if text not in role_map: return bot.send_message(uid, "⚠️ Tugmalardan birini bosing.")
    role, name = role_map[text], sessions[uid]["name"]
    sessions.pop(uid, None)
    data = load()
    data["pending"][str(uid)] = {"uid": uid, "name": name, "role": role, "username": msg.from_user.username or "", "time": now_str()}
    save(data)
    bot.send_message(uid, "✅ So'rovingiz yuborildi!\n\n⏳ Admin tasdiqlashini kuting.", reply_markup=types.ReplyKeyboardRemove())
    kb = types.InlineKeyboardMarkup()
    kb.add(types.InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"approve_{uid}_{role}"), types.InlineKeyboardButton("❌ Rad etish", callback_data=f"reject_{uid}"))
    notify_text = f"🆕 <b>Yangi so'rov</b>\n\n👤 {name}\n🆔 {uid}\n🎭 {ROLES[role]}\n📅 {now_str()}"
    bot.send_message(ADMIN_ID, notify_text, parse_mode="HTML", reply_markup=kb)
    # Agent bo'lsa supervayzrlarga ham yuborish
    if role == "agent":
        d2 = load()
        for u_id, u in d2["users"].items():
            if u.get("role") == "supervisor" and u.get("approved"):
                try:
                    kb2 = types.InlineKeyboardMarkup()
                    kb2.add(types.InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"approve_{uid}_{role}"),
                            types.InlineKeyboardButton("❌ Rad etish", callback_data=f"reject_{uid}"))
                    bot.send_message(int(u_id), notify_text, parse_mode="HTML", reply_markup=kb2)
                except: pass

@bot.callback_query_handler(func=lambda c: c.data.startswith("approve_"))
def approve_user(call):
    _, uid, role = call.data.split("_"); uid = int(uid)
    data = load(); p = data["pending"].get(str(uid))
    if not p: return bot.answer_callback_query(call.id, "Allaqachon ko'rib chiqilgan!")
    data["users"][str(uid)] = {"name": p["name"], "role": role, "username": p.get("username",""), "registered": p["time"], "approved": True, "total_visits": 0}
    data["pending"].pop(str(uid), None); save(data)
    bot.answer_callback_query(call.id, "✅ Tasdiqlandi!")
    bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
    bot.send_message(call.message.chat.id, f"✅ <b>{p['name']}</b> tasdiqlandi!", parse_mode="HTML")
    try: bot.send_message(uid, f"🎉 Tabriklaymiz, <b>{p['name']}</b>!\n\nXodim: {ROLES[role]}", parse_mode="HTML", reply_markup=main_kb(role))
    except: pass

@bot.callback_query_handler(func=lambda c: c.data.startswith("reject_"))
def reject_user(call):
    uid = int(call.data.split("_")[1]); data = load(); p = data["pending"].get(str(uid))
    if not p: return bot.answer_callback_query(call.id, "Topilmadi!")
    data["pending"].pop(str(uid), None); save(data)
    bot.answer_callback_query(call.id, "❌"); bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id)
    bot.send_message(call.message.chat.id, f"❌ <b>{p['name']}</b> rad etildi.", parse_mode="HTML")
    try: bot.send_message(uid, "❌ So'rovingiz rad etildi.")
    except: pass

# =============================================
# ⬅️ ORQAGA
# =============================================
@bot.message_handler(func=lambda m: m.text == "⬅️ Orqaga")
def go_back(msg):
    uid = msg.from_user.id; sess = sessions.get(uid, {}); user = get_user(uid)
    if not sess or sess.get("step") is None:
        if user and user.get("approved"):
            kb = admin_kb() if is_admin(uid) else main_kb(user["role"])
            return bot.send_message(uid, "↩️", reply_markup=kb)
        return
    step = sess.get("step", "")
    if step in ("location", "shop_select", "new_shop", "client_select"):
        sessions.pop(uid, None)
        kb = admin_kb() if is_admin(uid) else (main_kb(user["role"]) if user else types.ReplyKeyboardRemove())
        bot.send_message(uid, "↩️ Bekor qilindi.", reply_markup=kb)
    elif step == "photo":
        sess["step"] = "location"
        bot.send_message(uid, "📍 <b>1-qadam: Lokatsiya</b>", parse_mode="HTML", reply_markup=location_kb())
    elif step == "products":
        sess["step"] = "photo"
        bot.send_message(uid, "📸 <b>2-qadam: Foto</b>", parse_mode="HTML", reply_markup=back_kb())
    elif step == "vozvrat":
        sess["step"] = "products"
        bot.send_message(uid, "📦 <b>3-qadam: Mahsulotlar</b>", parse_mode="HTML", reply_markup=products_kb(sess["report"].get("product_counts", {})))
    elif step == "qty_sale":
        sess.pop("qty_input", None)
        sess["step"] = "products"
        bot.send_message(uid, "↩️", reply_markup=products_kb(sess["report"].get("product_counts", {})))
    elif step == "qty_vozvrat":
        sess.pop("qty_input", None)
        sess["step"] = "vozvrat"
        bot.send_message(uid, "↩️", reply_markup=vozvrat_kb(sess["report"].get("vozvrat_counts", {})))
    elif step == "polka_photo":
        sess["step"] = "vozvrat"
        bot.send_message(uid, "🔄 <b>4-qadam: Vozvratlar</b>\n\nQaytarilgan mahsulotlar:", parse_mode="HTML", reply_markup=vozvrat_kb(sess["report"].get("vozvrat_counts", {})))
    elif step in ("client_menu","client_search","client_add_name","client_add_address","client_add_phone","client_add_location","client_detail"):
        sessions.pop(uid, None)
        kb = admin_kb() if is_admin(uid) else (main_kb(user["role"]) if user else types.ReplyKeyboardRemove())
        bot.send_message(uid, "↩️", reply_markup=kb)
    elif step in ("my_report_menu","admin_report_menu","pick_agent","admin_date_from","admin_date_to","my_date_from","my_date_to","pick_product","pick_shop"):
        sessions.pop(uid, None)
        kb = admin_kb() if is_admin(uid) else (main_kb(user["role"]) if user else types.ReplyKeyboardRemove())
        bot.send_message(uid, "↩️", reply_markup=kb)

# =============================================
# 👥 MIJOZLAR BO'LIMI
# =============================================
@bot.message_handler(func=lambda m: m.text == "👥 Mijozlar")
def clients_main(msg):
    uid = msg.from_user.id
    if is_admin(uid):
        return show_client_menu(uid)
    user = get_user(uid)
    if not user or not user.get("approved"): return bot.send_message(uid, "❌ Ruxsat yo'q.")
    show_client_menu(uid)

def show_client_menu(uid):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add("🔍 Mijoz qidirish", "➕ Yangi mijoz")
    kb.add("📋 Barcha mijozlar")
    kb.add("📥 Excel import")
    kb.add("⬅️ Orqaga")
    sessions[uid] = {"step": "client_menu"}
    bot.send_message(uid, "👥 <b>Mijozlar bo'limi</b>\n\nTanlang:", parse_mode="HTML", reply_markup=kb)

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id, {}).get("step") == "client_menu")
def client_menu_handler(msg):
    uid, text = msg.from_user.id, msg.text
    if text == "⬅️ Orqaga":
        sessions.pop(uid, None)
        if is_admin(uid):
            return bot.send_message(uid, "↩️", reply_markup=admin_kb())
        user = get_user(uid)
        kb = main_kb(user["role"]) if user else types.ReplyKeyboardRemove()
        return bot.send_message(uid, "↩️", reply_markup=kb)
    elif text == "📋 Barcha mijozlar":
        show_all_clients(uid)
    elif text == "🔍 Mijoz qidirish":
        sessions[uid]["step"] = "client_search"
        bot.send_message(uid, "🔍 <b>Mijoz qidirish</b>\n\nMijoz nomi yoki adresini yozing:", parse_mode="HTML", reply_markup=back_kb())
    elif text == "➕ Yangi mijoz":
        start_add_client(uid)
    elif text == "📥 Excel import":
        sessions[uid]["step"] = "client_excel_import"
        bot.send_message(uid,
            "📥 <b>Excel import</b>\n\n"
            "Excel (.xlsx) faylini yuboring.\n\n"
            "📋 <b>Fayl formati:</b>\n"
            "A ustun: Mijoz nomi\n"
            "B ustun: Manzil\n"
            "C ustun: Telefon\n\n"
            "<i>1-qator sarlavha bo\'lishi mumkin — avtomatik o\'tkaziladi</i>",
            parse_mode="HTML", reply_markup=back_kb())


# --- Excel import ---
@bot.message_handler(content_types=["document"], func=lambda m: sessions.get(m.from_user.id, {}).get("step") == "client_excel_import")
def client_excel_import(msg):
    uid = msg.from_user.id
    if not msg.document:
        return bot.send_message(uid, "⚠️ Fayl yuboring!")
    fname = msg.document.file_name or ""
    if not fname.endswith(".xlsx"):
        return bot.send_message(uid, "❌ Faqat .xlsx fayl qabul qilinadi!")
    bot.send_message(uid, "⏳ Fayl o'qilmoqda...")
    try:
        import tempfile, openpyxl
        file_info = bot.get_file(msg.document.file_id)
        file_bytes = bot.download_file(file_info.file_path)
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp.write(file_bytes); tmp.close()
        wb = openpyxl.load_workbook(tmp.name)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        os.unlink(tmp.name)
        # 1-qator sarlavhami tekshir
        start_row = 0
        if rows and rows[0]:
            first = str(rows[0][0] or "").lower()
            if any(w in first for w in ["nom", "name", "mijoz", "client"]):
                start_row = 1
        clients = load_clients()
        existing_names = {c.get("name","").lower() for c in clients}
        added, skipped = 0, 0
        for row in rows[start_row:]:
            if not row or not row[0]: continue
            name    = str(row[0]).strip() if row[0] else ""
            address = str(row[1]).strip() if len(row) > 1 and row[1] else "—"
            phone   = str(row[2]).strip() if len(row) > 2 and row[2] else "—"
            if not name: continue
            if name.lower() in existing_names:
                skipped += 1; continue
            client = {
                "id": len(clients) + 1,
                "name": name,
                "address": address,
                "phone": phone,
                "lat": None,
                "lon": None,
                "created": now_str(),
                "source": "excel"
            }
            clients.append(client)
            existing_names.add(name.lower())
            added += 1
        save_clients(clients)
        sessions[uid]["step"] = "client_menu"
        bot.send_message(uid,
            f"✅ <b>Import yakunlandi!</b>\n\n"
            f"➕ Qo\'shildi: <b>{added} ta</b>\n"
            f"⏭ O\'tkazildi (avval bor): <b>{skipped} ta</b>\n"
            f"📋 Jami mijozlar: <b>{len(clients)} ta</b>",
            parse_mode="HTML")
        show_client_menu(uid)
    except Exception as e:
        print(f"❌ Excel import xato: {e}")
        bot.send_message(uid, f"❌ Xatolik: <code>{e}</code>\n\nFayl formatini tekshiring.", parse_mode="HTML")
        show_client_menu(uid)

# --- Barcha mijozlar ---
def show_all_clients(uid, page=0):
    clients = load_clients()
    if not clients: return bot.send_message(uid, "👥 Hali mijoz yo'q.\n➕ Yangi mijoz qo'shing!")
    per_page = 10; total = len(clients); pages = (total + per_page - 1) // per_page
    start, end = page * per_page, min((page + 1) * per_page, total)
    text = f"👥 <b>Mijozlar</b> ({start+1}-{end} / {total})\n\n"
    for c in clients[start:end]:
        text += f"  🆔 {c['id']}. <b>{c['name']}</b>\n"
        text += f"     📍 {c.get('address','—')}\n"
        text += f"     📞 {c.get('phone','—')}\n\n"
    kb = types.InlineKeyboardMarkup()
    if page > 0: kb.add(types.InlineKeyboardButton("◀️ Oldingi", callback_data=f"clients_page_{page-1}"))
    if end < total: kb.add(types.InlineKeyboardButton("Keyingi ▶️", callback_data=f"clients_page_{page+1}"))
    kb.add(types.InlineKeyboardButton("📍 Lokatsiyasini ko'rish", callback_data=f"client_map_select"))
    bot.send_message(uid, text, parse_mode="HTML", reply_markup=kb)

@bot.callback_query_handler(func=lambda c: c.data.startswith("clients_page_"))
def clients_page(call):
    page = int(call.data.split("_")[2])
    show_all_clients(call.from_user.id, page)
    bot.answer_callback_query(call.id)

# --- Qidiruv ---
@bot.message_handler(func=lambda m: sessions.get(m.from_user.id, {}).get("step") == "client_search")
def client_search(msg):
    uid, query = msg.from_user.id, msg.text.strip()
    if query == "⬅️ Orqaga": return show_client_menu(uid)
    results = search_clients(query)
    if not results:
        return bot.send_message(uid, f"🔍 <b>'{query}'</b> bo'yicha hech narsa topilmadi.\n\nBoshqa nom yozing:", parse_mode="HTML", reply_markup=back_kb())
    text = f"🔍 <b>'{query}'</b> bo'yicha natijalar: <b>{len(results)} ta</b>\n\n"
    for c in results:
        text += f"  🆔 {c['id']}. <b>{c['name']}</b>\n"
        text += f"     📍 {c.get('address','—')}\n"
        text += f"     📞 {c.get('phone','—')}\n\n"
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add("🔍 Boshqa qidirish")
    if any(c.get("lat") for c in results):
        kb.add("📍 Lokatsiyasini ko'rish")
    kb.add("📋 Barcha mijozlar", "⬅️ Orqaga")
    sessions[uid]["search_results"] = results
    sessions[uid]["step"] = "client_search_result"
    bot.send_message(uid, text, parse_mode="HTML", reply_markup=kb)

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id, {}).get("step") == "client_search_result")
def client_search_result_handler(msg):
    uid, text = msg.from_user.id, msg.text
    if text == "⬅️ Orqaga": return show_client_menu(uid)
    if text == "📋 Barcha mijozlar": return show_all_clients(uid)
    if text == "🔍 Boshqa qidirish":
        sessions[uid]["step"] = "client_search"
        return bot.send_message(uid, "🔍 Mijoz nomi yoki adresini yozing:", parse_mode="HTML", reply_markup=back_kb())
    if text == "📍 Lokatsiyasini ko'rish":
        results = sessions[uid].get("search_results", [])
        with_loc = [c for c in results if c.get("lat")]
        if not with_loc: return bot.send_message(uid, "📍 Bu mijozlar uchun lokatsiya kiritilmagan.")
        for c in with_loc:
            bot.send_location(uid, c["lat"], c["lon"])
            bot.send_message(uid, f"📍 <b>{c['name']}</b>\n{c.get('address','')}", parse_mode="HTML")
        return
    # Mijoz ID sini qidirish
    for c in sessions[uid].get("search_results", []):
        if text.strip() == str(c["id"]) or text.strip() == c["name"]:
            show_client_detail(uid, c)
            return
    bot.send_message(uid, "⚠️ Mijozni tanlang yoki tugmalardan foydalaning.")

# --- Mijoz kartasi ---
def show_client_detail(uid, client):
    sessions[uid] = {"step": "client_detail", "client": client}
    text = (
        "━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 <b>Mijoz kartasi</b>\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        f"🆔 ID: {client['id']}\n"
        f"👤 Nomi: <b>{client['name']}</b>\n"
        f"📍 Adres: {client.get('address','—')}\n"
        f"📞 Tel: {client.get('phone','—')}\n"
        f"📅 Qo'shilgan: {client.get('created','—')}\n"
        "━━━━━━━━━━━━━━━━━━━━"
    )
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    if client.get("lat"):
        kb.add("📍 Lokatsiyani ko'rish")
    kb.add("🔍 Boshqa qidirish", "⬅️ Orqaga")
    bot.send_message(uid, text, parse_mode="HTML", reply_markup=kb)

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id, {}).get("step") == "client_detail")
def client_detail_handler(msg):
    uid, text = msg.from_user.id, msg.text
    if text == "⬅️ Orqaga": return show_client_menu(uid)
    if text == "🔍 Boshqa qidirish":
        sessions[uid]["step"] = "client_search"
        return bot.send_message(uid, "🔍 Mijoz nomi yoki adresini yozing:", parse_mode="HTML", reply_markup=back_kb())
    if text == "📍 Lokatsiyani ko'rish":
        client = sessions[uid].get("client", {})
        if client.get("lat"):
            bot.send_location(uid, client["lat"], client["lon"])
            bot.send_message(uid, f"📍 <b>{client['name']}</b>\n{client.get('address','')}", parse_mode="HTML")
        else:
            bot.send_message(uid, "📍 Lokatsiya kiritilmagan.")
        return
    show_client_menu(uid)

# --- Yangi mijoz qo'shish ---
def start_add_client(uid):
    sessions[uid] = {"step": "client_add_name", "new_client": {}}
    bot.send_message(uid, "➕ <b>Yangi mijoz</b>\n\n1️⃣ Mijoz nomini yozing:", parse_mode="HTML", reply_markup=back_kb())

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id, {}).get("step") == "client_add_name")
def client_add_name(msg):
    uid, text = msg.from_user.id, msg.text.strip()
    if text == "⬅️ Orqaga": return show_client_menu(uid)
    if len(text) < 2: return bot.send_message(uid, "⚠️ Kamida 2 belgi yozing!")
    sessions[uid]["new_client"]["name"] = text
    sessions[uid]["step"] = "client_add_address"
    bot.send_message(uid, f"✅ Nomi: <b>{text}</b>\n\n2️⃣ Adresini yozing:", parse_mode="HTML", reply_markup=back_kb())

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id, {}).get("step") == "client_add_address")
def client_add_address(msg):
    uid, text = msg.from_user.id, msg.text.strip()
    if text == "⬅️ Orqaga":
        sessions[uid]["step"] = "client_add_name"
        return bot.send_message(uid, "1️⃣ Mijoz nomini yozing:", parse_mode="HTML", reply_markup=back_kb())
    sessions[uid]["new_client"]["address"] = text
    sessions[uid]["step"] = "client_add_phone"
    bot.send_message(uid, f"✅ Adres: <b>{text}</b>\n\n3️⃣ Tel raqamini yozing:", parse_mode="HTML", reply_markup=back_kb())

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id, {}).get("step") == "client_add_phone")
def client_add_phone(msg):
    uid, text = msg.from_user.id, msg.text.strip()
    if text == "⬅️ Orqaga":
        sessions[uid]["step"] = "client_add_address"
        return bot.send_message(uid, "2️⃣ Adresini yozing:", parse_mode="HTML", reply_markup=back_kb())
    sessions[uid]["new_client"]["phone"] = text
    sessions[uid]["step"] = "client_add_location"
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    kb.add(types.KeyboardButton("📍 Lokatsiyani yuborish", request_location=True))
    kb.add("⏭ O'tkazib yuborish", "⬅️ Orqaga")
    bot.send_message(uid, f"✅ Tel: <b>{text}</b>\n\n4️⃣ Lokatsiyani yuboring yoki o'tkazib yuboring:", parse_mode="HTML", reply_markup=kb)

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id, {}).get("step") == "client_add_location")
def client_add_location(msg):
    uid = msg.from_user.id
    if msg.text == "⬅️ Orqaga":
        sessions[uid]["step"] = "client_add_phone"
        return bot.send_message(uid, "3️⃣ Tel raqamini yozing:", parse_mode="HTML", reply_markup=back_kb())
    if msg.text == "⏭ O'tkazib yuborish":
        nc = sessions[uid]["new_client"]
        client = add_client(nc["name"], nc["address"], nc["phone"])
        sessions.pop(uid, None)
        bot.send_message(uid, f"✅ Mijoz saqlandi!\n\n👤 <b>{client['name']}</b>\n📍 {client['address']}\n📞 {client['phone']}", parse_mode="HTML")
        return show_client_menu(uid)
    if msg.location:
        nc = sessions[uid]["new_client"]
        client = add_client(nc["name"], nc["address"], nc["phone"], msg.location.latitude, msg.location.longitude)
        sessions.pop(uid, None)
        bot.send_location(uid, client["lat"], client["lon"])
        bot.send_message(uid, f"✅ Mijoz lokatsiya bilan saqlandi!\n\n👤 <b>{client['name']}</b>\n📍 {client['address']}\n📞 {client['phone']}", parse_mode="HTML")
        return show_client_menu(uid)
    bot.send_message(uid, "⚠️ Lokatsiya yuboring yoki 'O'tkazib yuborish' bosing.")

# --- Lokatsiyani ko'rish (callback) ---
@bot.callback_query_handler(func=lambda c: c.data == "client_map_select")
def client_map_select(call):
    uid = call.from_user.id
    bot.answer_callback_query(call.id)
    bot.send_message(uid, "📍 Qaysi mijozning lokatsiyasini ko'rmoqchisiz?\nMijoz ID sini yoki nomini yozing:", reply_markup=back_kb())
    sessions[uid] = {"step": "client_map_input"}

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id, {}).get("step") == "client_map_input")
def client_map_input(msg):
    uid, text = msg.from_user.id, msg.text.strip()
    if text == "⬅️ Orqaga": return show_client_menu(uid)
    clients = load_clients()
    found = None
    for c in clients:
        if str(c.get("id")) == text or c.get("name","").lower() == text.lower():
            found = c; break
    if not found: return bot.send_message(uid, "❌ Topilmadi. Boshqa ID yoki nom yozing:", reply_markup=back_kb())
    if not found.get("lat"): return bot.send_message(uid, f"📍 <b>{found['name']}</b> uchun lokatsiya kiritilmagan.", parse_mode="HTML")
    bot.send_location(uid, found["lat"], found["lon"])
    bot.send_message(uid, f"📍 <b>{found['name']}</b>\n{found.get('address','')}", parse_mode="HTML")
    sessions.pop(uid, None)
    show_client_menu(uid)

# =============================================
# 🛒 SAVDONI BOSHLASH
# =============================================
@bot.message_handler(func=lambda m: m.text == "🛒 Savdoni boshlash")
def start_visit(msg):
    uid = msg.from_user.id; user = get_user(uid)
    if not user or not user.get("approved") or user["role"] != "agent":
        return bot.send_message(uid, "❌ Ruxsat yo'q.")
    sessions[uid] = {"step": "location", "report": {"agent_id": uid, "agent_name": user["name"], "started": now_str(), "date": today_str()}}
    bot.send_message(uid, "📍 <b>1-qadam: Lokatsiya</b>\n\nHozirgi joylashuvingizni yuboring:", parse_mode="HTML", reply_markup=location_kb())

@bot.message_handler(content_types=["location"])
def receive_location(msg):
    uid = msg.from_user.id; sess = sessions.get(uid)
    if not sess: return

    # --- Mijoz lokatsiyasi qo'shish ---
    if sess.get("step") == "client_add_location":
        nc = sess.get("new_client", {})
        client = add_client(nc.get("name",""), nc.get("address",""), nc.get("phone",""),
                            msg.location.latitude, msg.location.longitude)
        sessions.pop(uid, None)
        bot.send_location(uid, client["lat"], client["lon"])
        bot.send_message(uid,
            f"✅ Mijoz lokatsiya bilan saqlandi!\n\n"
            f"👤 <b>{client['name']}</b>\n"
            f"📍 {client['address']}\n"
            f"📞 {client['phone']}",
            parse_mode="HTML")
        return show_client_menu(uid)

    # --- Savdo lokatsiyasi ---
    if sess.get("step") != "location": return
    lat, lon = msg.location.latitude, msg.location.longitude
    sess["report"]["location"] = {"lat": lat, "lon": lon}

    nearby_clients = find_nearby_clients_gps(lat, lon, 300)
    nearby_shops   = find_nearby_shops(lat, lon, 300)

    if nearby_clients:
        # Klientlarni saqlash (orqaga bosganda kerak bo'ladi)
        sess["step"] = "client_select"
        sess["nearby_clients"] = nearby_clients[:15]
        sess["nearby_shops"]   = nearby_shops
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
        kb.add("➕ Yangi mijoz")
        for c in nearby_clients[:15]:
            kb.add(f"👤 {c['name']} ({c['distance']}m)")
        kb.add("⬅️ Orqaga")
        bot.send_message(uid,
            "✅ Lokatsiya qabul qilindi!\n\n"
            "👥 <b>300m ichidagi klientlar:</b>\nBirini tanlang yoki yangi qo'shing:",
            parse_mode="HTML", reply_markup=kb)
    elif nearby_shops:
        sess["step"] = "shop_select"; sess["nearby_shops"] = nearby_shops
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
        kb.add("➕ Yangi mijoz")
        for s in nearby_shops: kb.add(f"🏪 {s['name']} ({s['distance']}m)")
        kb.add("⬅️ Orqaga")
        bot.send_message(uid,
            "✅ Lokatsiya qabul qilindi!\n\n📡 300m ichidagi magazinlar:\nTanlang:",
            parse_mode="HTML", reply_markup=kb)
    else:
        sess["step"] = "new_shop"
        bot.send_message(uid,
            "✅ Lokatsiya qabul qilindi!\n\n"
            "📡 Yaqinda hech kim topilmadi.\n🆕 <b>Mijoz nomini yozing:</b>",
            parse_mode="HTML", reply_markup=back_kb())

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id, {}).get("step") == "client_select")
def select_client(msg):
    uid = msg.from_user.id; sess = sessions.get(uid)
    if not sess: return
    text = (msg.text or "").strip()

    if text == "⬅️ Orqaga":
        sessions.pop(uid, None)
        user = get_user(uid)
        kb = admin_kb() if is_admin(uid) else (main_kb(user["role"]) if user else types.ReplyKeyboardRemove())
        return bot.send_message(uid, "↩️ Bekor qilindi.", reply_markup=kb)

    if text == "➕ Yangi mijoz":
        nearby_shops = sess.get("nearby_shops", [])
        if nearby_shops:
            sess["step"] = "shop_select"
            kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
            for s in nearby_shops: kb.add(f"🏪 {s['name']} ({s['distance']}m)")
            kb.add("🆕 Yangi magazin", "⬅️ Orqaga")
            return bot.send_message(uid, "🏪 Magazinni tanlang:", reply_markup=kb)
        else:
            sess["step"] = "new_shop"
            return bot.send_message(uid, "🆕 Mijoz nomini yozing:", reply_markup=back_kb())

    # Tanlangan klientni topish
    for c in sess.get("nearby_clients", []):
        if text == f"👤 {c['name']} ({c['distance']}m)":
            sess["report"]["shop_name"]    = c["name"]
            sess["report"]["shop_id"]      = None
            sess["report"]["client_name"]  = c["name"]
            sess["nearby_clients_bak"]     = sess.get("nearby_clients", [])
            sess.pop("nearby_clients", None)
            sess["step"] = "photo"
            return bot.send_message(uid,
                f"✅ Mijoz: <b>{c['name']}</b>\n"
                f"📍 {c.get('address', '—')}\n"
                f"📏 {c['distance']} metr\n\n"
                "📸 <b>2-qadam: Foto</b>\n\nJonli foto yuboring:",
                parse_mode="HTML", reply_markup=back_kb())

    bot.send_message(uid, "⚠️ Ro'yxatdan tanlang yoki ➕ Yangi mijoz bosing.")

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id, {}).get("step") == "shop_select")
def select_shop(msg):
    uid = msg.from_user.id
    sess = sessions.get(uid)
    if not sess: return
    text = (msg.text or "").strip()
    if text == "⬅️ Orqaga":
        sessions.pop(uid, None)
        user = get_user(uid)
        return bot.send_message(uid, "↩️", reply_markup=main_kb(user["role"]) if user else types.ReplyKeyboardRemove())
    if text == "🆕 Yangi magazin":
        sess["step"] = "new_shop"
        return bot.send_message(uid, "🏪 Yangi magazin nomini yozing:", reply_markup=back_kb())
    for s in sess.get("nearby_shops", []):
        if text.startswith(f"🏪 {s['name']}"):
            sess["report"]["shop_name"], sess["report"]["shop_id"] = s["name"], s["id"]
            sess.pop("nearby_shops", None); sess["step"] = "photo"
            return bot.send_message(uid, f"✅ {s['name']}\n\n📸 <b>2-qadam: Foto</b>", parse_mode="HTML", reply_markup=back_kb())
    bot.send_message(uid, "⚠️ Ro'yxatdan tanlang.")
@bot.message_handler(func=lambda m: sessions.get(m.from_user.id, {}).get("step") == "new_shop")
def new_shop_name(msg):
    uid = msg.from_user.id
    sess = sessions.get(uid, {})
    text = msg.text.strip()

    if text == "⬅️ Orqaga":
        # Oldingi yoki location qadamiga qaytish
        lat = sess.get("report", {}).get("location", {}).get("lat")
        lon = sess.get("report", {}).get("location", {}).get("lon")
        if lat and lon:
            nearby = find_nearby_shops(lat, lon, 100)
            if nearby:
                sess["step"] = "shop_select"
                sess["nearby_shops"] = nearby
                kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
                for s in nearby:
                    kb.add(f"🏪 {s['name']} ({s['distance']}m)")
                kb.add("🆕 Yangi magazin", "⬅️ Orqaga")
                bot.send_message(uid, "📡 100m ichidagi magazinlar:", reply_markup=kb)
            else:
                bot.send_message(uid, "🏪 Yangi magazin nomini yozing:", reply_markup=back_kb())
        else:
            # Lokatsiya yo'q bo'lsa butunlay bekor qil
            sessions.pop(uid, None)
            user = get_user(uid)
            kb = admin_kb() if is_admin(uid) else main_kb(user.get("role", "agent"))
            bot.send_message(uid, "↩️ Bekor qilindi.", reply_markup=kb)
        return

    if len(text) < 2:
        return bot.send_message(uid, "⚠️ Magazin nomi kamida 2 belgi boʻlishi kerak. Qaytadan yozing:")

    lat = sess.get("report", {}).get("location", {}).get("lat")
    lon = sess.get("report", {}).get("location", {}).get("lon")
    if not lat or not lon:
        # Agar lokatsiya yo'qolgan bo'lsa, qaytadan lokatsiya so'rang
        sess["step"] = "location"
        return bot.send_message(uid, "📍 Avval lokatsiyani yuboring!", reply_markup=location_kb())

    try:
        shop = add_shop(text, lat, lon)
        sess["report"]["shop_name"] = shop["name"]
        sess["report"]["shop_id"] = shop["id"]
        sess["step"] = "photo"
        bot.send_message(uid,
            f"✅ Yangi magazin: <b>{shop['name']}</b>\n\n"
            "📸 <b>2-qadam: Foto</b>\n\n"
            "📷 Kamerani oching va jonli foto yuboring:",
            parse_mode="HTML", reply_markup=back_kb())
    except Exception as e:
        print(f"XATOLIK: {e}")
        bot.send_message(uid,
            f"❌ Magazinni saqlashda xatolik:\n<code>{e}</code>\n\n"
            "Iltimos, qayta urinib koʻring yoki adminga xabar bering.",
            parse_mode="HTML", reply_markup=back_kb())

@bot.message_handler(content_types=["photo"], func=lambda m: sessions.get(m.from_user.id, {}).get("step") == "photo")
def receive_photo(msg):
    uid = msg.from_user.id; sess = sessions.get(uid)
    if not sess or sess["step"] != "photo": return
    if msg.forward_date or msg.media_group_id: return bot.send_message(uid, "❌ Faqat kameradan jonli foto!")
    if int(time.time()) - msg.date > 120: return bot.send_message(uid, "❌ Eski foto!")
    sess["report"]["photo_id"], sess["report"]["photo_time"] = msg.photo[-1].file_id, now_str()
    sess["step"], sess["report"]["product_counts"] = "products", {}
    pl = "\n".join(f"  {i+1}. {p['name']} — {fmt(p['price'])} so'm" for i, p in enumerate(PRODUCTS))
    bot.send_message(uid, f"✅ Foto qabul qilindi!\n\n📦 <b>3-qadam: Mahsulotlar</b>\n\n{pl}", parse_mode="HTML", reply_markup=products_kb({}))

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id, {}).get("step") == "products")
def select_product(msg):
    uid = msg.from_user.id
    sess = sessions.get(uid)
    if not sess: return
    text = (msg.text or "").strip()
    if text == "✅ Tayyor":
        if not sess["report"]["product_counts"]:
            return bot.send_message(uid, "⚠️ Kamida 1 ta mahsulot tanlang!")
        sess["step"] = "vozvrat"
        sess["report"]["vozvrat_counts"] = {}
        return bot.send_message(uid,
            "🔴 <b>4-qadam: Vozvratlar</b>\n\n"
            "Qaytarilgan mahsulotlarni belgilang.\n"
            "Vozvrat yo'q bo'lsa ✅ Tayyor bosing:",
            parse_mode="HTML", reply_markup=vozvrat_kb({}))
    if text == "⬅️ Orqaga":
        return  # go_back handler hal qiladi
    # Mahsulot tanlov — indeks bo'yicha xavfsiz moslashtirish
    for i, p in enumerate(PRODUCTS):
        cnt = sess["report"]["product_counts"].get(str(i), 0)
        label_with_cnt = f"{p['name']} [{cnt} ta]"
        # Ikki variantni ham tekshiramiz: sof nom va sonli nom
        if text == p['name'] or text == label_with_cnt:
            sess["step"] = "qty_sale"
            sess["selected_product"] = i
            sess["qty_input"] = ""
            cur_qty = sess["report"]["product_counts"].get(str(i), 0)
            msg_text = (
                f"📦 <b>{p['name']}</b>\n"
                f"Narxi: <b>{fmt(p['price'])} so'm</b>\n"
                + (f"Joriy soni: <b>{cur_qty} ta</b>\n" if cur_qty > 0 else "")
                + "\nYangi soni:"
            )
            return bot.send_message(uid, msg_text, parse_mode="HTML", reply_markup=qty_numpad_kb())

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id, {}).get("step") == "qty_sale")
def receive_qty_sale(msg):
    uid = msg.from_user.id
    sess = sessions.get(uid)
    if not sess: return
    text = (msg.text or "").strip()
    idx, p = sess["selected_product"], PRODUCTS[sess["selected_product"]]
    if text.isdigit():
        cur = sess.get("qty_input","") + text; cur = "0" if cur.lstrip("0")=="" else str(int(cur))
        sess["qty_input"] = cur; qty = int(cur)
        bot.send_message(uid, f"📦 <b>{p['name']}</b>\nSoni: <b>{cur}</b> ta" + (f"\nJami: <b>{fmt(qty*p['price'])} so'm</b>" if qty>0 else ""), parse_mode="HTML", reply_markup=qty_numpad_kb())
    elif text == "🔙":
        cur = sess.get("qty_input",""); cur = cur[:-1] if len(cur)>1 else "0"; sess["qty_input"] = cur
        bot.send_message(uid, f"📦 <b>{p['name']}</b>\nSoni: <b>{cur}</b> ta", parse_mode="HTML", reply_markup=qty_numpad_kb())
    elif text == "✅ Tasdiqlash":
        qty = int(sess.get("qty_input","0")); sess.pop("qty_input",None)
        if qty>0: sess["report"]["product_counts"][str(idx)] = qty
        else: sess["report"]["product_counts"].pop(str(idx), None)
        sess["step"] = "products"
        bot.send_message(uid, f"✅ <b>{p['name']}</b> — {qty} ta" + (f" ({fmt(qty*p['price'])} so'm)" if qty>0 else ""), parse_mode="HTML", reply_markup=products_kb(sess["report"]["product_counts"]))
    elif text == "⬅️ Orqaga": sess.pop("qty_input",None); sess["step"]="products"; bot.send_message(uid, "↩️", reply_markup=products_kb(sess["report"]["product_counts"]))
    else: bot.send_message(uid, "⚠️ Raqamlardan foydalaning!")

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id, {}).get("step") == "vozvrat")
def select_vozvrat(msg):
    uid = msg.from_user.id
    sess = sessions.get(uid)
    if not sess: return
    text = (msg.text or "").strip()
    # 🔴 prefixini olib tashlash
    clean_text = text.replace("🔴 ", "")
    if text == "✅ Tayyor":
        sess["report"]["expired"] = []
        # 5-qadam: polka foto
        sess["step"] = "polka_photo"
        return bot.send_message(uid,
            "📸 <b>5-qadam: Polka foto</b>\n\n"
            "Mahsulotni joylashtirgan polkangiz rasmini yuboring:",
            parse_mode="HTML", reply_markup=back_kb())
    if text == "⬅️ Orqaga":
        return  # go_back handler hal qiladi
    for i, p in enumerate(PRODUCTS):
        cnt = sess["report"]["vozvrat_counts"].get(str(i), 0)
        label_with_cnt = f"{p['name']} [{cnt} ta]"
        if clean_text == p['name'] or clean_text == label_with_cnt:
            sess["step"] = "qty_vozvrat"
            sess["selected_product"] = i
            sess["qty_input"] = ""
            cur_qty = sess["report"]["vozvrat_counts"].get(str(i), 0)
            msg_text = (
                f"🔴 <b>Vozvrat: {p['name']}</b>\n"
                + (f"Joriy vozvrat: <b>{cur_qty} ta</b>\n" if cur_qty > 0 else "")
                + "\nYangi soni:"
            )
            return bot.send_message(uid, msg_text, parse_mode="HTML", reply_markup=qty_numpad_vozv_kb())

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id, {}).get("step") == "qty_vozvrat")
def receive_qty_vozvrat(msg):
    uid = msg.from_user.id
    sess = sessions.get(uid)
    if not sess: return
    text = (msg.text or "").strip()
    idx, p = sess["selected_product"], PRODUCTS[sess["selected_product"]]
    if text.isdigit():
        cur = sess.get("qty_input","")+text; cur = "0" if cur.lstrip("0")=="" else str(int(cur))
        sess["qty_input"] = cur; bot.send_message(uid, f"🔴 <b>Vozvrat: {p['name']}</b>\nSoni: <b>{cur}</b> ta", parse_mode="HTML", reply_markup=qty_numpad_vozv_kb())
    elif text == "🔙":
        cur = sess.get("qty_input",""); cur = cur[:-1] if len(cur)>1 else "0"; sess["qty_input"]=cur
        bot.send_message(uid, f"🔴 <b>Vozvrat: {p['name']}</b>\nSoni: <b>{cur}</b> ta", parse_mode="HTML", reply_markup=qty_numpad_vozv_kb())
    elif text == "✅ Tasdiqlash":
        qty = int(sess.get("qty_input","0")); sess.pop("qty_input",None)
        if qty>0: sess["report"]["vozvrat_counts"][str(idx)] = qty
        else: sess["report"]["vozvrat_counts"].pop(str(idx), None)
        sess["step"]="vozvrat"; bot.send_message(uid, f"🔴 <b>{p['name']}</b> vozvrat — {qty} ta", parse_mode="HTML", reply_markup=vozvrat_kb(sess["report"]["vozvrat_counts"]))
    elif text == "⬅️ Orqaga": sess.pop("qty_input",None); sess["step"]="vozvrat"; bot.send_message(uid, "↩️", reply_markup=vozvrat_kb(sess["report"]["vozvrat_counts"]))
    else: bot.send_message(uid, "⚠️ Raqamlardan foydalaning!")

@bot.message_handler(content_types=["photo"], func=lambda m: sessions.get(m.from_user.id, {}).get("step") == "polka_photo")
def receive_polka_photo(msg):
    uid = msg.from_user.id; sess = sessions.get(uid)
    if not sess: return
    if msg.forward_date:
        return bot.send_message(uid, "❌ Forward foto yuborildi! Kameradan yangi foto oling.")
    sess["report"]["polka_photo_id"] = msg.photo[-1].file_id
    sess["report"]["polka_photo_time"] = now_str()
    try:
        finish_report(uid)
    except Exception as e:
        print(f"❌ finish_report xatosi: {e}")
        u = get_user(uid)
        bot.send_message(uid, "⚠️ Texnik xato yuz berdi. Admin bilan bog'laning.", reply_markup=main_kb(u["role"]) if u else types.ReplyKeyboardRemove())
        sessions.pop(uid, None)

def auto_save_client_from_report(r):
    """Savdo yakunlanganda do'konni mijozlar ro'yxatiga avtomatik qo'shish."""
    shop_name = r.get("shop_name", "").strip()
    if not shop_name:
        return
    clients = load_clients()
    # Agar bu do'kon avval mijozlar ro'yxatida bo'lsa, qayta qo'shmaymiz
    for c in clients:
        if c.get("name", "").lower() == shop_name.lower():
            return
    lat = r.get("location", {}).get("lat")
    lon = r.get("location", {}).get("lon")
    client = {
        "id": len(clients) + 1,
        "name": shop_name,
        "address": "—",
        "phone": "—",
        "lat": lat,
        "lon": lon,
        "created": now_str(),
        "source": "savdo"  # qo'lda emas, avtomatik qo'shilgan
    }
    clients.append(client)
    save_clients(clients)

def finish_report(uid):
    sess = sessions.get(uid); r = sess["report"]; r["finished"] = now_str()
    sl, ts = [], 0
    for i, q in r.get("product_counts",{}).items(): s=q*PRODUCTS[int(i)]["price"]; ts+=s; sl.append(f"  • {PRODUCTS[int(i)]['name']}: {q} ta = {fmt(s)} so'm")
    vl, tv = [], 0
    for i, q in r.get("vozvrat_counts",{}).items(): s=q*PRODUCTS[int(i)]["price"]; tv+=s; vl.append(f"  • {PRODUCTS[int(i)]['name']}: {q} ta = {fmt(s)} so'm")
    nt = ts-tv; r["total_sale"], r["total_vozv"], r["net_total"] = ts, tv, nt
    d=load(); d["reports"].append(r)
    if str(uid) in d["users"]: d["users"][str(uid)]["total_visits"] = d["users"][str(uid)].get("total_visits",0)+1
    save(d)
    # ✅ Do'konni avtomatik mijozlar ro'yxatiga qo'shish
    auto_save_client_from_report(r)
    st="\n".join(sl) or "  —"; vt="\n".join(vl) or "  —"
    sm = f"━━━━━━━━━━━━━━━━━━━━\n✅ <b>Hisobot yakunlandi!</b>\n━━━━━━━━━━━━━━━━━━━━\n🏪 {r.get('shop_name','')}\n🕐 {r['started']} → {r['finished']}\n\n📦 Sotuv:\n{st}\n💰 Sotuv jami: <b>{fmt(ts)} so'm</b>\n\n🔄 Vozvrat:\n{vt}\n↩️ Vozvrat jami: <b>{fmt(tv)} so'm</b>\n\n💵 Sof: <b>{fmt(nt)} so'm</b>\n━━━━━━━━━━━━━━━━━━━━"
    bot.send_message(uid, sm, parse_mode="HTML", reply_markup=main_kb(get_user(uid)["role"]))
    at = f"━━━━━━━━━━━━━━━━━━━━\n📋 <b>Yangi hisobot</b>\n━━━━━━━━━━━━━━━━━━━━\n👤 {r['agent_name']}\n🏪 {r.get('shop_name','')}\n🕐 {r['started']} → {r['finished']}\n\n📦 Sotuv:\n{st}\n💰 Sotuv: <b>{fmt(ts)} so'm</b>\n\n🔄 Vozvrat:\n{vt}\n↩️ Vozvrat: <b>{fmt(tv)} so'm</b>\n\n💵 Sof: <b>{fmt(nt)} so'm</b>\n━━━━━━━━━━━━━━━━━━━━"
    bot.send_photo(ADMIN_ID, r["photo_id"], caption=at, parse_mode="HTML")
    if r.get("polka_photo_id"):
        bot.send_photo(ADMIN_ID, r["polka_photo_id"], caption="🛒 Polka foto")
    bot.send_location(ADMIN_ID, r["location"]["lat"], r["location"]["lon"])
    d2=load()
    for u_id,u in d2["users"].items():
        if u.get("role") in ("supervisor","manager") and u.get("approved"):
            try:
                bot.send_photo(int(u_id), r["photo_id"], caption=at, parse_mode="HTML")
                if r.get("polka_photo_id"):
                    bot.send_photo(int(u_id), r["polka_photo_id"], caption="🛒 Polka foto")
                bot.send_location(int(u_id), r["location"]["lat"], r["location"]["lon"])
            except: pass
    sessions.pop(uid, None)

# =============================================
# 📊 HISOBOT TIZIMI (qisqartirilgan)
# =============================================
def get_reports_by_period(sd, ed, aid=None, pi=None, sn=None):
    res=[]
    for r in load()["reports"]:
        rd=r.get("date",r.get("started","")[:10])
        if sd<=rd<=ed:
            if aid and str(r.get("agent_id"))!=str(aid): continue
            if pi is not None and int(r.get("product_counts",{}).get(str(pi),0))==0: continue
            if sn and r.get("shop_name","").lower()!=sn.lower(): continue
            res.append(r)
    return res

def calc_stats(reps):
    if not reps: return None
    ts=tv=0; ag={}; pr={i:{"name":p["name"],"price":p["price"],"qty":0,"sum":0} for i,p in enumerate(PRODUCTS)}
    vz={i:{"name":p["name"],"qty":0,"sum":0} for i,p in enumerate(PRODUCTS)}; sh={}
    for r in reps:
        ts+=r.get("total_sale",0); tv+=r.get("total_vozv",0)
        an=r.get("agent_name","?"); ag[an]=ag.get(an,{"visits":0,"sale":0,"vozv":0,"net":0}); ag[an]["visits"]+=1; ag[an]["sale"]+=r.get("total_sale",0); ag[an]["vozv"]+=r.get("total_vozv",0); ag[an]["net"]+=r.get("net_total",0)
        for i,q in r.get("product_counts",{}).items(): ii=int(i); pr[ii]["qty"]+=q; pr[ii]["sum"]+=q*pr[ii]["price"]
        for i,q in r.get("vozvrat_counts",{}).items(): ii=int(i); vz[ii]["qty"]+=q; vz[ii]["sum"]+=q*PRODUCTS[ii]["price"]
        sn=r.get("shop_name","?"); sh[sn]=sh.get(sn,{"visits":0,"sale":0,"vozv":0,"net":0}); sh[sn]["visits"]+=1; sh[sn]["sale"]+=r.get("total_sale",0); sh[sn]["vozv"]+=r.get("total_vozv",0); sh[sn]["net"]+=r.get("net_total",0)
    return {"total_visits":len(reps),"total_agents":len(ag),"total_sale":ts,"total_vozv":tv,"net_total":ts-tv,"agents":ag,"products":pr,"vozvrat":vz,"shops":sh,"reps":reps}

def fmt_report(stats, title=""):
    if not stats: return f"📭 {title}: ma'lumot yo'q."
    t=f"━━━━━━━━━━━━━━━━━━━━\n📊 <b>{title}</b>\n━━━━━━━━━━━━━━━━━━━━\n👥 Agentlar: <b>{stats['total_agents']}</b>\n🏪 Visit: <b>{stats['total_visits']}</b>\n🏬 Magazin: <b>{len(stats['shops'])}</b>\n\n📦 <b>Mahsulotlar:</b>\n"
    for p in stats["products"].values():
        if p["qty"]>0: t+=f"  • {p['name']}: {p['qty']} ta — {fmt(p['sum'])} so'm\n"
    if all(p["qty"]==0 for p in stats["products"].values()): t+="  —\n"
    if any(v["qty"]>0 for v in stats["vozvrat"].values()):
        t+="\n🔄 <b>Vozvrat:</b>\n"
        for v in stats["vozvrat"].values():
            if v["qty"]>0: t+=f"  • {v['name']}: {v['qty']} ta — {fmt(v['sum'])} so'm\n"
    t+=f"\n💰 Sotuv: <b>{fmt(stats['total_sale'])} so'm</b>\n↩️ Vozvrat: <b>{fmt(stats['total_vozv'])} so'm</b>\n💵 Sof: <b>{fmt(stats['net_total'])} so'm</b>\n━━━━━━━━━━━━━━━━━━━━"
    if stats["agents"]:
        sa=sorted(stats["agents"].items(),key=lambda x:x[1]["net"],reverse=True)
        t+="\n\n👤 <b>TOP Agentlar:</b>\n"
        for i,(n,a) in enumerate(sa[:10]): t+=f"  {i+1}. {n} — {a['visits']} Visit, {fmt(a['net'])} so'm\n"
    if stats["shops"]:
        ss=sorted(stats["shops"].items(),key=lambda x:x[1]["net"],reverse=True)
        t+="\n🏪 <b>TOP Magazinlar:</b>\n"
        for i,(n,s) in enumerate(ss[:10]): t+=f"  {i+1}. {n} — {s['visits']} Visit, {fmt(s['net'])} so'm\n"
    return t

def gen_excel(reps, fn="hisobot.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    C_DARK  = "1F3864"; C_HEAD  = "2E75B6"; C_HEAD2 = "17375E"
    C_WHITE = "FFFFFF"; C_GREEN = "E2EFDA"; C_RED   = "FCE4D6"
    C_BLUE  = "DEEAF1"; C_ALT   = "F2F7FB"; C_TOTAL = "BDD7EE"
    C_SUMBOX= "D6E4F0"; C_DARK2 = "1F3864"

    thin = Side(style="thin", color="C0C0C0")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    SOM  = "#,##0"

    def h(ws, row, col, val, bg=C_HEAD, fg=C_WHITE, sz=10, bold=True, align="center"):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(bold=bold, color=fg, name="Calibri", size=sz)
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        c.border    = brd
        return c

    def d(ws, row, col, val, fmt=None, bg=None, bold=False, align="center"):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name="Calibri", size=10, bold=bold)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border    = brd
        if fmt: c.number_format = fmt
        if bg:  c.fill = PatternFill("solid", start_color=bg)
        return c

    # ── 1-VARAQ: Batafsil ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Batafsil hisobot"
    ws.sheet_view.showGridLines = False

    last_col = 5 + len(PRODUCTS)*2 + 3

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws["A1"]
    c.value     = "SAVDO HISOBOTI  —  Agent nazorati MEHR"
    c.font      = Font(bold=True, size=15, color=C_WHITE, name="Calibri")
    c.fill      = PatternFill("solid", start_color=C_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    total_sale = sum(r.get("total_sale",0) for r in reps)
    total_vozv = sum(r.get("total_vozv",0) for r in reps)
    net_total  = sum(r.get("net_total", 0) for r in reps)
    agents_set = set(r.get("agent_name","") for r in reps)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    c = ws["A2"]
    c.value     = (f"Sana: {datetime.now(UZ_TZ).strftime('%Y-%m-%d')}   |   Visit: {len(reps)}   |   "
                   f"Agentlar: {len(agents_set)}   |   "
                   f"Sotuv: {total_sale:,} so'm   |   "
                   f"Vozvrat: {total_vozv:,} so'm   |   "
                   f"Sof: {net_total:,} so'm")
    c.font      = Font(size=10, color=C_DARK, name="Calibri", bold=True)
    c.fill      = PatternFill("solid", start_color=C_SUMBOX)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 6

    ROW_H = 4
    base_h = ["#", "Sana", "Vaqt", "Agent", "Mijoz / Do'kon"]
    prod_h = []
    for p in PRODUCTS:
        pn = p['name'].replace("🥪","").replace("🧒","").replace("🍗","").replace("🍔","").strip()
        prod_h += [f"{pn}\n(dona)", f"{pn}\n(so'm)"]
    sum_h  = ["Sotuv\n(so'm)", "Vozvrat\n(so'm)", "SOF\n(so'm)"]
    all_h  = base_h + prod_h + sum_h

    for col, hdr in enumerate(all_h, 1):
        bg = C_HEAD2 if col > len(base_h)+len(prod_h) else C_HEAD
        h(ws, ROW_H, col, hdr, bg=bg)
    ws.row_dimensions[ROW_H].height = 42

    for idx, r in enumerate(reps, 1):
        row    = ROW_H + idx
        bg_row = C_ALT if idx % 2 == 0 else None
        tstr   = r.get("started","")[11:16]
        d(ws, row, 1, idx,                    bg=bg_row)
        d(ws, row, 2, r.get("date",""),       bg=bg_row)
        d(ws, row, 3, tstr,                   bg=bg_row)
        d(ws, row, 4, r.get("agent_name",""), bg=bg_row, align="left")
        d(ws, row, 5, r.get("shop_name",""),  bg=bg_row, align="left")
        col = 6
        for i, p in enumerate(PRODUCTS):
            q  = r.get("product_counts",{}).get(str(i), 0)
            bg = C_BLUE if q > 0 else bg_row
            d(ws, row, col,   q,            bg=bg);        col += 1
            d(ws, row, col,   q*p["price"], SOM, bg=bg);  col += 1
        d(ws, row, col, r.get("total_sale",0), SOM, bg=C_BLUE if r.get("total_sale",0)>0 else bg_row); col+=1
        d(ws, row, col, r.get("total_vozv",0), SOM, bg=C_RED  if r.get("total_vozv",0)>0 else bg_row); col+=1
        d(ws, row, col, r.get("net_total", 0), SOM, bg=C_GREEN, bold=True)
        ws.row_dimensions[row].height = 18

    if reps:
        tr = ROW_H + len(reps) + 1
        ws.row_dimensions[tr].height = 24
        for col, val in enumerate(["","","","JAMI", f"{len(reps)} visit"], 1):
            h(ws, tr, col, val, bg=C_TOTAL, fg=C_DARK)
        col = 6; ds = ROW_H+1; de = ROW_H+len(reps)
        for i in range(len(PRODUCTS)*2):
            cl = get_column_letter(col+i)
            c  = ws.cell(tr, col+i, f"=SUM({cl}{ds}:{cl}{de})")
            c.font = Font(bold=True, name="Calibri", size=10, color=C_DARK)
            c.fill = PatternFill("solid", start_color=C_TOTAL)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = brd
            if i%2==1: c.number_format = SOM
        col += len(PRODUCTS)*2
        for offset, bgc in enumerate([C_BLUE, C_RED, C_GREEN]):
            cl = get_column_letter(col+offset)
            c  = ws.cell(tr, col+offset, f"=SUM({cl}{ds}:{cl}{de})")
            c.font = Font(bold=True, name="Calibri", size=11, color=C_DARK)
            c.fill = PatternFill("solid", start_color=bgc)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = brd; c.number_format = SOM

    widths = [4, 12, 8, 20, 28]
    for _ in PRODUCTS: widths += [8, 16]
    widths += [16, 16, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"F{ROW_H+1}"
    ws.auto_filter.ref = f"A{ROW_H}:{get_column_letter(len(all_h))}{ROW_H}"

    # ── 2-VARAQ: Agent reytingi ───────────────────────────────────────
    ws2 = wb.create_sheet("Agent reytingi")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:H1")
    c = ws2["A1"]
    c.value     = "AGENT REYTINGI  —  Agent nazorati MEHR"
    c.font      = Font(bold=True, size=14, color=C_WHITE, name="Calibri")
    c.fill      = PatternFill("solid", start_color=C_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 34
    ws2.row_dimensions[2].height = 6

    for col, hdr in enumerate(["#","Agent","Visit","Sotuv (so'm)","Vozvrat (so'm)","Sof (so'm)","O'rtacha (so'm)","Ulushi %"], 1):
        h(ws2, 3, col, hdr, bg=C_HEAD2)
    ws2.row_dimensions[3].height = 36

    agent_stats = {}
    for r in reps:
        an = r.get("agent_name","Noma'lum")
        if an not in agent_stats:
            agent_stats[an] = {"visits":0,"sale":0,"vozv":0,"net":0}
        agent_stats[an]["visits"] += 1
        agent_stats[an]["sale"]   += r.get("total_sale",0)
        agent_stats[an]["vozv"]   += r.get("total_vozv",0)
        agent_stats[an]["net"]    += r.get("net_total", 0)

    medals = ["🥇","🥈","🥉"]
    for idx, (name, st) in enumerate(sorted(agent_stats.items(), key=lambda x:x[1]["net"], reverse=True), 1):
        row    = idx + 3
        bg_row = C_ALT if idx%2==0 else None
        avg    = st["sale"]//st["visits"] if st["visits"] else 0
        share  = round(st["net"]/net_total*100,1) if net_total else 0
        d(ws2, row, 1, medals[idx-1] if idx<=3 else str(idx), bg=bg_row, bold=(idx<=3))
        d(ws2, row, 2, name,       bg=bg_row, align="left", bold=(idx==1))
        d(ws2, row, 3, st["visits"],bg=bg_row)
        d(ws2, row, 4, st["sale"], SOM, bg=C_BLUE)
        d(ws2, row, 5, st["vozv"], SOM, bg=C_RED  if st["vozv"]>0 else bg_row)
        d(ws2, row, 6, st["net"],  SOM, bg=C_GREEN, bold=True)
        d(ws2, row, 7, avg,        SOM, bg=bg_row)
        d(ws2, row, 8, f"{share}%",bg=bg_row)
        ws2.row_dimensions[row].height = 20

    if agent_stats:
        tr2 = len(agent_stats)+4
        ws2.row_dimensions[tr2].height = 24
        h(ws2,tr2,1,"JAMI",bg=C_TOTAL,fg=C_DARK)
        h(ws2,tr2,2,f"{len(agent_stats)} agent",bg=C_TOTAL,fg=C_DARK)
        h(ws2,tr2,3,len(reps),bg=C_TOTAL,fg=C_DARK)
        for col,val,bgc in [(4,total_sale,C_BLUE),(5,total_vozv,C_RED),(6,net_total,C_GREEN)]:
            c=ws2.cell(tr2,col,val); c.number_format=SOM
            c.font=Font(bold=True,name="Calibri",size=10,color=C_DARK)
            c.fill=PatternFill("solid",start_color=bgc)
            c.alignment=Alignment(horizontal="center",vertical="center"); c.border=brd
        h(ws2,tr2,7,"",bg=C_TOTAL,fg=C_DARK)
        h(ws2,tr2,8,"100%",bg=C_TOTAL,fg=C_DARK)

    for col,w in zip("ABCDEFGH",[6,24,8,18,18,20,20,12]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A4"

    path = f"/tmp/{fn}"
    wb.save(path)
    return path


def send_report_menu(uid):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add("📅 Bugun", "📆 Bu oy")
    user = get_user(uid)
    if user and user.get("role") != "agent":
        kb.add("👤 Agent bo'yicha", "📦 Mahsulot bo'yicha")
    else:
        kb.add("📦 Mahsulot bo'yicha")
    kb.add("🏪 Magazin bo'yicha", "🗓 Sana oralig'i")
    kb.add("📤 Excel yuklash", "📋 Batafsil hisobot")
    kb.add("⬅️ Orqaga")
    bot.send_message(uid, "📊 <b>Hisobotlar paneli</b>", parse_mode="HTML", reply_markup=kb)

@bot.message_handler(func=lambda m: m.text == "📊 Mening hisobotlarim")
def my_reports(msg):
    uid=msg.from_user.id
    if get_user(uid).get("role")!="agent": return
    sessions[uid]={"step":"my_report_menu","agent_filter":uid}; send_report_menu(uid)

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id,{}).get("step")=="my_report_menu")
def my_rep(msg):
    uid,us=msg.from_user.id,get_user(msg.from_user.id); tx=msg.text; ai=sessions[uid].get("agent_filter",uid)
    if tx=="⬅️ Orqaga": sessions.pop(uid,None); return bot.send_message(uid,"↩️",reply_markup=main_kb(us["role"]) if us else types.ReplyKeyboardRemove())
    if tx=="📅 Bugun": reps=get_reports_by_period(today_str(),today_str(),ai); bot.send_message(uid,fmt_report(calc_stats(reps),"Bugun"),parse_mode="HTML")
    elif tx=="📆 Bu oy": reps=get_reports_by_period(datetime.now(UZ_TZ).strftime("%Y-%m-01"),today_str(),ai); bot.send_message(uid,fmt_report(calc_stats(reps),"Bu oy"),parse_mode="HTML")
    elif tx=="📤 Excel yuklash": reps=get_reports_by_period("2000-01-01",today_str(),ai); p=gen_excel(reps); bot.send_document(uid,open(p,"rb"),caption=f"📤 Hisobot Excel | {len(reps)} ta yozuv")
    elif tx=="📦 Mahsulot bo'yicha": show_product_filter(uid,"my_report_menu")
    elif tx=="🏪 Magazin bo'yicha": show_shop_filter(uid,"my_report_menu")
    elif tx=="📋 Batafsil hisobot": reps=get_reports_by_period("2000-01-01",today_str(),ai); bot.send_message(uid,fmt_report(calc_stats(reps),"Barcha hisobotlarim"),parse_mode="HTML")
    elif tx=="🗓 Sana oralig'i": sessions[uid]["step"]="my_date_from"; bot.send_message(uid,"Boshlanish sanasi:\n<i>2026-04-01</i>",parse_mode="HTML",reply_markup=back_kb())

def is_mgr(uid):
    if is_admin(uid): return True
    u = get_user(uid)
    return bool(u and u.get("role") in ("supervisor","manager") and u.get("approved"))

@bot.message_handler(func=lambda m: m.text=="📊 Hisobotlar" and is_mgr(m.from_user.id))
def admin_reports(msg):
    uid=msg.from_user.id; sessions[uid]={"step":"admin_report_menu","agent_filter":None}; send_report_menu(uid)

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id,{}).get("step")=="admin_report_menu")
def adm_rep(msg):
    uid=msg.from_user.id; us=get_user(uid); tx=msg.text; ai=sessions[uid].get("agent_filter")
    rl="admin" if is_admin(uid) else (us["role"] if us else "agent")
    if tx=="⬅️ Orqaga": sessions.pop(uid,None); return bot.send_message(uid,"↩️",reply_markup=admin_kb() if is_admin(uid) else main_kb(rl))
    if tx=="📅 Bugun": reps=get_reports_by_period(today_str(),today_str(),ai); bot.send_message(uid,fmt_report(calc_stats(reps),"Bugun"),parse_mode="HTML")
    elif tx=="📆 Bu oy": reps=get_reports_by_period(datetime.now(UZ_TZ).strftime("%Y-%m-01"),today_str(),ai); bot.send_message(uid,fmt_report(calc_stats(reps),"Bu oy"),parse_mode="HTML")
    elif tx=="📤 Excel yuklash": reps=get_reports_by_period("2000-01-01",today_str(),ai); p=gen_excel(reps); bot.send_document(uid,open(p,"rb"),caption=f"📤 Hisobot Excel | {len(reps)} ta yozuv")
    elif tx=="👤 Agent bo'yicha": show_agent_filter(uid,"admin_report_menu")
    elif tx=="📦 Mahsulot bo'yicha": show_product_filter(uid,"admin_report_menu")
    elif tx=="🏪 Magazin bo'yicha": show_shop_filter(uid,"admin_report_menu")
    elif tx=="📋 Batafsil hisobot": reps=get_reports_by_period("2000-01-01",today_str(),ai); bot.send_message(uid,fmt_report(calc_stats(reps),"Barcha"),parse_mode="HTML")
    elif tx=="🗓 Sana oralig'i": sessions[uid]["step"]="admin_date_from"; bot.send_message(uid,"Boshlanish sanasi:\n<i>2026-04-01</i>",parse_mode="HTML",reply_markup=back_kb())

def show_agent_filter(uid,rs):
    agents=[(u_id,u) for u_id,u in load()["users"].items() if u.get("role")=="agent"]
    if not agents: return bot.send_message(uid,"Hali agent yo'q.")
    kb=types.ReplyKeyboardMarkup(resize_keyboard=True); kb.add("🌐 Barcha agentlar")
    for u_id,u in agents: kb.add(f"{u['name']} [{u_id}]")
    kb.add("⬅️ Orqaga")
    prev_filter = sessions.get(uid, {}).get("agent_filter")
    sessions[uid]={"step":"pick_agent","agent_return":rs,"agent_filter":prev_filter,"agents_map":{f"{u['name']} [{u_id}]":u_id for u_id,u in agents}}
    bot.send_message(uid,"👤 Agentni tanlang:",reply_markup=kb)

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id,{}).get("step")=="pick_agent")
def pick_agent(msg):
    uid,tx=msg.from_user.id,msg.text; rs=sessions[uid].get("agent_return","admin_report_menu")
    if tx=="⬅️ Orqaga": sessions[uid]={"step":rs,"agent_filter":None}; return send_report_menu(uid)
    if tx=="🌐 Barcha agentlar": sessions[uid]["agent_filter"]=None
    else:
        am=sessions[uid].get("agents_map",{})
        if tx in am: sessions[uid]["agent_filter"]=am[tx]
        else: return bot.send_message(uid,"⚠️ Agentni tanlang.")
    sessions[uid]["step"]=rs
    bot.send_message(uid,f"✅ Agent: <b>{tx}</b>",parse_mode="HTML"); send_report_menu(uid)

def show_product_filter(uid,rs):
    kb=types.ReplyKeyboardMarkup(resize_keyboard=True); kb.add("📦 Barcha mahsulotlar")
    for p in PRODUCTS: kb.add(p['name'])
    kb.add("⬅️ Orqaga")
    prev_filter = sessions.get(uid, {}).get("agent_filter")
    sessions[uid]={"step":"pick_product","product_return":rs,"agent_filter":prev_filter}
    bot.send_message(uid,"📦 Mahsulotni tanlang:",reply_markup=kb)

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id,{}).get("step")=="pick_product")
def pick_product(msg):
    uid,tx=msg.from_user.id,msg.text; rs=sessions[uid].get("product_return","admin_report_menu")
    if tx=="⬅️ Orqaga": sessions[uid]={"step":rs,"agent_filter":sessions[uid].get("agent_filter")}; return send_report_menu(uid)
    pi=None
    if tx!="📦 Barcha mahsulotlar":
        for i,p in enumerate(PRODUCTS):
            if tx==p['name']: pi=i; break
        if pi is None: return bot.send_message(uid,"⚠️ Mahsulotni tanlang.")
    ai=sessions[uid].get("agent_filter"); reps=get_reports_by_period("2000-01-01",today_str(),ai,pi)
    bot.send_message(uid,fmt_report(calc_stats(reps),f"Mahsulot: {tx}"),parse_mode="HTML")
    sessions[uid]={"step":rs,"agent_filter":ai}; send_report_menu(uid)

def show_shop_filter(uid,rs):
    sn=set()
    for r in load()["reports"]:
        if r.get("shop_name"): sn.add(r["shop_name"])
    kb=types.ReplyKeyboardMarkup(resize_keyboard=True); kb.add("🏬 Barcha magazinlar")
    for n in sorted(sn): kb.add(f"🏪 {n}")
    kb.add("⬅️ Orqaga")
    prev_filter = sessions.get(uid, {}).get("agent_filter")
    sessions[uid]={"step":"pick_shop","shop_return":rs,"agent_filter":prev_filter}
    bot.send_message(uid,"🏪 Magazinni tanlang:",reply_markup=kb)

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id,{}).get("step")=="pick_shop")
def pick_shop(msg):
    uid,tx=msg.from_user.id,msg.text; rs=sessions[uid].get("shop_return","admin_report_menu")
    if tx=="⬅️ Orqaga": sessions[uid]={"step":rs,"agent_filter":sessions[uid].get("agent_filter")}; return send_report_menu(uid)
    sn=None
    if tx!="🏬 Barcha magazinlar": sn=tx.replace("🏪 ","")
    ai=sessions[uid].get("agent_filter"); reps=get_reports_by_period("2000-01-01",today_str(),ai,sn=sn)
    bot.send_message(uid,fmt_report(calc_stats(reps),f"Magazin: {sn or 'Barcha'}"),parse_mode="HTML")
    sessions[uid]={"step":rs,"agent_filter":ai}; send_report_menu(uid)

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id,{}).get("step") in ("my_date_from","admin_date_from"))
def date_from(msg):
    uid,ss=msg.from_user.id,sessions[uid]; tx=msg.text.strip()
    if tx=="⬅️ Orqaga": rs="my_report_menu" if ss["step"]=="my_date_from" else "admin_report_menu"; sessions[uid]={"step":rs,"agent_filter":ss.get("agent_filter")}; return send_report_menu(uid)
    try:
        datetime.strptime(tx,"%Y-%m-%d"); ss["date_from"]=tx
        ss["step"]="my_date_to" if "my" in ss["step"] else "admin_date_to"
        bot.send_message(uid,"Tugash sanasi:\n<i>2026-04-30</i>",parse_mode="HTML")
    except: bot.send_message(uid,"❌ Format: 2026-04-01")

@bot.message_handler(func=lambda m: sessions.get(m.from_user.id,{}).get("step") in ("my_date_to","admin_date_to"))
def date_to(msg):
    uid,ss=msg.from_user.id,sessions[uid]; tx=msg.text.strip()
    if tx=="⬅️ Orqaga": rs="my_report_menu" if ss["step"]=="my_date_to" else "admin_report_menu"; sessions[uid]={"step":rs,"agent_filter":ss.get("agent_filter")}; return send_report_menu(uid)
    try:
        datetime.strptime(tx,"%Y-%m-%d"); rs="my_report_menu" if ss["step"]=="my_date_to" else "admin_report_menu"
        df,dt=ss["date_from"],tx; sessions[uid]["step"]=rs
        reps=get_reports_by_period(df,dt,ss.get("agent_filter"))
        bot.send_message(uid,fmt_report(calc_stats(reps),f"{df} → {dt}"),parse_mode="HTML")
        send_report_menu(uid)
    except: bot.send_message(uid,"❌ Format: 2026-04-30")

# =============================================
# QOLGAN BO'LIMLAR
# =============================================
@bot.message_handler(func=lambda m: m.text=="👥 Foydalanuvchilar" and is_admin(m.from_user.id))
def admin_users(msg):
    d=load(); t="👥 <b>Foydalanuvchilar:</b>\n\n"
    for uid,u in d["users"].items():
        if str(uid)==str(ADMIN_ID): continue
        t+=f"{ROLES.get(u['role'],'')} <b>{u['name']}</b>\n   🆔 {uid} | 🏪 {u.get('total_visits',0)} Visit\n\n"
    bot.send_message(ADMIN_ID,t,parse_mode="HTML")

@bot.message_handler(func=lambda m: m.text=="👥 Agentlar" and is_mgr(m.from_user.id))
def view_agents(msg):
    uid=msg.from_user.id; d=load()
    agents=[(u_id,u) for u_id,u in d["users"].items() if u.get("role")=="agent" and u.get("approved")]
    if not agents: return bot.send_message(uid,"Hali agent yo'q.")
    t="👥 <b>Agentlar:</b>\n\n"
    for u_id,u in agents:
        reps=[r for r in d["reports"] if str(r.get("agent_id"))==u_id]
        td=[r for r in reps if r.get("date")==today_str()]
        t+=f"👤 <b>{u['name']}</b>\n   🏪 Jami: {len(reps)} | 📅 Bugun: {len(td)}\n\n"
    bot.send_message(uid,t,parse_mode="HTML")

@bot.message_handler(func=lambda m: m.text=="🏆 Reyting" and is_mgr(m.from_user.id))
def rating(msg):
    uid=msg.from_user.id; d=load(); sc=[]
    for u_id,u in d["users"].items():
        if u.get("role")!="agent": continue
        reps=[r for r in d["reports"] if str(r.get("agent_id"))==u_id]
        sc.append((u["name"],len(reps),sum(r.get("net_total",0) for r in reps)))
    sc.sort(key=lambda x:x[2],reverse=True)
    md=["🥇","🥈","🥉"]; t="🏆 <b>Reyting</b>\n\n"
    for i,(n,v,tot) in enumerate(sc): ic=md[i] if i<3 else f"{i+1}."; t+=f"{ic} <b>{n}</b>\n   🏪 {v} Visit | 💵 {fmt(tot)} so'm\n\n"
    bot.send_message(uid,t or "Ma'lumot yo'q.",parse_mode="HTML")

@bot.message_handler(func=lambda m: m.text=="💰 Sotuv statistikasi" and is_mgr(m.from_user.id))
def sales_stats(msg):
    uid=msg.from_user.id; rt=get_reports_by_period(today_str(),today_str()); rm=get_reports_by_period(datetime.now(UZ_TZ).strftime("%Y-%m-01"),today_str())
    stt=calc_stats(rt); stm=calc_stats(rm); t="💰 <b>Sotuv statistikasi</b>\n━━━━━━━━━━━━━━━━━━━━\n📅 <b>Bugun:</b>\n"
    if stt and stt["total_visits"]>0:
        t+=f"  Visit: {stt['total_visits']}\n"
        for p in stt["products"].values():
            if p["qty"]>0: t+=f"  • {p['name']}: {p['qty']} ta = {fmt(p['sum'])} so'm\n"
        t+=f"  💵 Sof: <b>{fmt(stt['net_total'])} so'm</b>\n"
    else: t+="  —\n"
    t+="\n📆 <b>Bu oy:</b>\n"
    if stm and stm["total_visits"]>0:
        t+=f"  Visit: {stm['total_visits']}\n"
        for p in stm["products"].values():
            if p["qty"]>0: t+=f"  • {p['name']}: {p['qty']} ta = {fmt(p['sum'])} so'm\n"
        t+=f"  💵 Sof: <b>{fmt(stm['net_total'])} so'm</b>\n"
    else: t+="  —\n"
    t+="━━━━━━━━━━━━━━━━━━━━"
    bot.send_message(uid,t,parse_mode="HTML")

@bot.message_handler(func=lambda m: m.text=="🏪 Magazinlar" and is_mgr(m.from_user.id))
def list_shops(msg):
    uid=msg.from_user.id; sh=load_shops()
    if not sh: return bot.send_message(uid,"🏪 Magazin yo'q.")
    t="🏪 <b>Magazinlar:</b>\n\n"
    for s in sh: t+=f"  🆔 {s['id']}. <b>{s['name']}</b>\n     📍 {s['lat']}, {s['lon']}\n\n"
    bot.send_message(uid,t,parse_mode="HTML")

@bot.message_handler(func=lambda m: m.text=="⏳ Kutayotganlar" and is_admin(m.from_user.id))
def pending_list(msg):
    d=load()
    if not d.get("pending"): return bot.send_message(ADMIN_ID,"⏳ Kutayotgan so'rov yo'q.")
    for u_id,p in d["pending"].items():
        kb=types.InlineKeyboardMarkup()
        kb.add(types.InlineKeyboardButton("✅",callback_data=f"approve_{u_id}_{p['role']}"),types.InlineKeyboardButton("❌",callback_data=f"reject_{u_id}"))
        bot.send_message(ADMIN_ID,f"⏳ {p['name']}\n🆔 {u_id}\n🎭 {ROLES.get(p['role'],'')}",parse_mode="HTML",reply_markup=kb)

# =============================================
# 📅 KUN YAKUNIY HTML HISOBOT (soat 20:00 da avtomatik)
# =============================================
def get_photo_base64(file_id: str) -> str:
    """Telegram file_id dan base64 rasm olish"""
    try:
        file_info = bot.get_file(file_id)
        file_bytes = bot.download_file(file_info.file_path)
        import base64
        return base64.b64encode(file_bytes).decode("utf-8")
    except Exception as e:
        print(f"⚠️ Rasm yuklab bo'lmadi: {e}")
        return ""

def gen_daily_html(target_date: str) -> str:
    d = load()
    reps = [r for r in d["reports"] if r.get("date", r.get("started","")[:10]) == target_date]
    agents = {}
    for r in reps:
        an = r.get("agent_name","Noma'lum")
        agents.setdefault(an, []).append(r)
    total_sale = sum(r.get("total_sale",0) for r in reps)
    total_vozv = sum(r.get("total_vozv",0) for r in reps)
    net_total  = sum(r.get("net_total", 0) for r in reps)
    agent_rows = ""
    for agent_name, visits in agents.items():
        a_sale = sum(v.get("total_sale",0) for v in visits)
        a_vozv = sum(v.get("total_vozv",0) for v in visits)
        a_net  = sum(v.get("net_total", 0) for v in visits)
        vrows = ""
        for i, v in enumerate(visits, 1):
            prods = [f"{PRODUCTS[int(idx)]['name']}: {qty} ta ({fmt(qty*PRODUCTS[int(idx)]['price'])} so'm)"
                     for idx,qty in v.get("product_counts",{}).items() if qty>0]
            vozv  = [f"{PRODUCTS[int(idx)]['name']}: {qty} ta"
                     for idx,qty in v.get("vozvrat_counts",{}).items() if qty>0]
            t1 = v.get("started","")[:16]; t2 = v.get("finished","")[:16]
            tstr = f"{t1} → {t2}" if t2 else t1

            # --- Rasmlar ---
            photo_b64 = get_photo_base64(v["photo_id"]) if v.get("photo_id") else ""
            polka_b64 = get_photo_base64(v["polka_photo_id"]) if v.get("polka_photo_id") else ""

            photo_html = (
                f'<div class="photos">'
                + (f'<div class="photo-wrap"><div class="photo-lbl">📸 Kirish foto</div>'
                   f'<img src="data:image/jpeg;base64,{photo_b64}" alt="foto"/></div>'
                   if photo_b64 else '')
                + (f'<div class="photo-wrap"><div class="photo-lbl">🛒 Polka foto</div>'
                   f'<img src="data:image/jpeg;base64,{polka_b64}" alt="polka foto"/></div>'
                   if polka_b64 else '')
                + '</div>'
            ) if (photo_b64 or polka_b64) else ''

            vrows += f"""<tr>
              <td class="num">{i}</td>
              <td>
                <b>{v.get('shop_name','—')}</b>
                {photo_html}
              </td>
              <td class="center">{tstr}</td>
              <td>{"<br>".join(prods) or "—"}</td>
              <td class="red">{"<br>".join(vozv) or "—"}</td>
              <td class="money">{fmt(v.get('total_sale',0))}</td>
              <td class="money red">{fmt(v.get('total_vozv',0))}</td>
              <td class="money green">{fmt(v.get('net_total',0))}</td>
            </tr>"""
        agent_rows += f"""<div class="agent-block">
          <div class="agent-header">👤 {agent_name}
            <span class="badge">{len(visits)} visit</span>
            <span class="badge green">{fmt(a_net)} so'm</span>
          </div>
          <div class="tbl-wrap"><table>
            <thead><tr><th>#</th><th>Mijoz/Magazin</th><th>Vaqt</th>
              <th>Sotuv</th><th>Vozvrat</th>
              <th>Sotuv ₸</th><th>Vozvrat ₸</th><th>Sof ₸</th></tr></thead>
            <tbody>{vrows}</tbody>
            <tfoot><tr><td colspan="5" class="right"><b>JAMI:</b></td>
              <td class="money"><b>{fmt(a_sale)}</b></td>
              <td class="money red"><b>{fmt(a_vozv)}</b></td>
              <td class="money green"><b>{fmt(a_net)}</b></td></tr></tfoot>
          </table></div></div>"""
    if not agent_rows:
        agent_rows = '<div class="empty">📭 Bu kun hech qanday hisobot yo\'q.</div>'
    return f"""<!DOCTYPE html><html lang="uz"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Hisobot {target_date}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',sans-serif;background:#f0f2f5;color:#1a1a2e}}
.header{{background:linear-gradient(135deg,#1a1a2e,#16213e);color:#fff;padding:24px;text-align:center}}
.header h1{{font-size:22px;margin-bottom:4px}}
.header p{{font-size:13px;opacity:.7}}
.summary{{display:flex;gap:12px;padding:16px;flex-wrap:wrap}}
.card{{flex:1;min-width:130px;background:#fff;border-radius:12px;padding:14px 18px;box-shadow:0 2px 8px rgba(0,0,0,.08);text-align:center}}
.card .val{{font-size:20px;font-weight:700;margin-bottom:4px}}
.card .lbl{{font-size:11px;color:#888;text-transform:uppercase}}
.card.green .val{{color:#27ae60}}.card.red .val{{color:#e74c3c}}.card.blue .val{{color:#2980b9}}
.content{{padding:0 16px 24px}}
.agent-block{{background:#fff;border-radius:14px;margin-bottom:20px;box-shadow:0 2px 10px rgba(0,0,0,.07);overflow:hidden}}
.agent-header{{background:#16213e;color:#fff;padding:12px 18px;font-size:15px;font-weight:600;display:flex;gap:10px;align-items:center;flex-wrap:wrap}}
.badge{{background:rgba(255,255,255,.2);border-radius:20px;padding:2px 10px;font-size:12px}}
.badge.green{{background:#27ae60}}
.tbl-wrap{{overflow-x:auto}}
table{{width:100%;border-collapse:collapse;font-size:13px}}
thead tr{{background:#f8f9fa}}
th{{padding:10px 12px;text-align:left;font-size:11px;text-transform:uppercase;color:#666;border-bottom:2px solid #eee}}
td{{padding:10px 12px;border-bottom:1px solid #f0f0f0;vertical-align:top}}
tfoot td{{background:#f8f9fa;border-top:2px solid #dee2e6}}
.num{{color:#aaa;font-size:12px;width:30px}}.center{{text-align:center;white-space:nowrap;color:#666}}
.right{{text-align:right}}.money{{text-align:right;white-space:nowrap;font-weight:600}}
.green{{color:#27ae60}}.red{{color:#e74c3c}}
.empty{{text-align:center;padding:40px;color:#aaa;font-size:15px}}
.photos{{display:flex;gap:8px;margin-top:8px;flex-wrap:wrap}}
.photo-wrap{{display:flex;flex-direction:column;align-items:center;gap:4px}}
.photo-wrap img{{width:120px;height:90px;object-fit:cover;border-radius:8px;border:2px solid #e0e0e0;box-shadow:0 1px 4px rgba(0,0,0,.12)}}
.photo-lbl{{font-size:10px;color:#888;text-align:center}}
@media(max-width:600px){{.summary{{flex-direction:column}}th,td{{padding:7px 8px;font-size:12px}}.photo-wrap img{{width:90px;height:68px}}}}
</style></head><body>
<div class="header">
  <h1>📊 Kunlik Hisobot</h1>
  <p>{target_date} — Agent nazorati MEHR</p>
  <p style="margin-top:6px;font-size:12px;opacity:.6">Yaratildi: {datetime.now(UZ_TZ).strftime('%H:%M')} (O'ZB vaqti)</p>
</div>
<div class="summary">
  <div class="card blue"><div class="val">{len(reps)}</div><div class="lbl">Jami visit</div></div>
  <div class="card blue"><div class="val">{len(agents)}</div><div class="lbl">Agentlar</div></div>
  <div class="card"><div class="val">{fmt(total_sale)}</div><div class="lbl">Sotuv (so'm)</div></div>
  <div class="card red"><div class="val">{fmt(total_vozv)}</div><div class="lbl">Vozvrat (so'm)</div></div>
  <div class="card green"><div class="val">{fmt(net_total)}</div><div class="lbl">Sof daromad</div></div>
</div>
<div class="content">{agent_rows}</div></body></html>"""


def send_daily_report(target_date: str = None):
    if not target_date:
        target_date = (datetime.now(UZ_TZ) - timedelta(days=1)).strftime("%Y-%m-%d")
    html = gen_daily_html(target_date)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".html", mode="w", encoding="utf-8")
    tmp.write(html); tmp.close()
    caption = f"📊 <b>Kunlik hisobot</b> — {target_date}"
    d = load()
    recipients = [ADMIN_ID]
    for u_id, u in d["users"].items():
        if u.get("role") in ("supervisor","manager") and u.get("approved"):
            try: recipients.append(int(u_id))
            except: pass
    for rec in set(recipients):
        try:
            with open(tmp.name,"rb") as f:
                bot.send_document(rec, f, caption=caption, parse_mode="HTML",
                                  visible_file_name=f"hisobot_{target_date}.html")
        except Exception as e:
            print(f"❌ {rec}: {e}")
    os.unlink(tmp.name)
    print(f"✅ Hisobot yuborildi: {target_date}")


def midnight_scheduler():
    while True:
        now = datetime.now(UZ_TZ)
        # Bugungi soat 20:00 O'zbekiston vaqti
        next_run = now.replace(hour=20, minute=0, second=0, microsecond=0)
        if next_run <= now:
            next_run += timedelta(days=1)
        wait = (next_run - now).total_seconds()
        print(f"⏳ Keyingi hisobot (O'ZB vaqti): {next_run.strftime('%Y-%m-%d %H:%M')} ({int(wait//3600)}s {int((wait%3600)//60)}m)")
        time.sleep(wait)
        # Bugungi kun hisobotini yuborish (soat 20:00 da)
        send_daily_report(datetime.now(UZ_TZ).strftime("%Y-%m-%d"))


# ─── Admin: qo'lda hisobot olish ───────────────────────────────────────
@bot.message_handler(func=lambda m: m.text=="📤 Kunlik hisobot" and is_admin(m.from_user.id))
def manual_daily_report(msg):
    bot.send_message(msg.from_user.id, "⏳ Hisobot tayyorlanmoqda...")
    send_daily_report(today_str())
    bot.send_message(msg.from_user.id, "✅ Bugungi hisobot yuborildi!")

# =============================================
@bot.message_handler(func=lambda m: m.text=="🔄 Hammasini restart" and is_admin(m.from_user.id))
def restart_all(msg):
    uid = msg.from_user.id
    count = len(sessions)
    sessions.clear()
    # Barcha foydalanuvchilarga xabar yuborish
    d = load()
    notified = 0
    for u_id, u in d["users"].items():
        if str(u_id) == str(ADMIN_ID): continue
        if u.get("approved"):
            try:
                bot.send_message(int(u_id),
                    "🔄 <b>Tizim qayta ishga tushirildi!</b>\n\n"
                    "Iltimos, /start bosing.",
                    parse_mode="HTML")
                notified += 1
            except:
                pass
    bot.send_message(uid,
        f"✅ <b>Restart bajarildi!</b>\n\n"
        f"🗑 Tozalangan sessiyalar: {count} ta\n"
        f"📨 Xabar yuborildi: {notified} ta xodim",
        parse_mode="HTML", reply_markup=admin_kb())

print("✅ Bot ishga tushdi...")

# ⏰ Scheduler — alohida thread da ishlaydi
sched_thread = threading.Thread(target=midnight_scheduler, daemon=True)
sched_thread.start()

bot.infinity_polling()